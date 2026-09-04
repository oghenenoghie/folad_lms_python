"""CSV/XLSX bulk import for Staff records — see
apps.students.services.bulk_import_service for the shared shape (this
mirrors it for Staff's own required fields rather than sharing code,
since the two models' required columns and validation differ enough that
a shared helper would need as much branching as just having two).
"""
import csv
import io

from django.db import IntegrityError, transaction

from apps.schools.models import School
from apps.staff.models import EMPLOYMENT_STATUS_CHOICES, Staff

from . import staff_service

REQUIRED_COLUMNS = ["school_code", "first_name", "last_name", "position", "date_joined"]
_EMPLOYMENT_STATUSES = {choice for choice, _ in EMPLOYMENT_STATUS_CHOICES}


class BulkImportError(ValueError):
    pass


def parse_rows(*, filename: str, content: bytes) -> list[dict]:
    if filename.lower().endswith(".xlsx"):
        return _parse_xlsx(content)
    return _parse_csv(content)


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows, [])]
    return [
        {header[i]: cell for i, cell in enumerate(row) if i < len(header)}
        for row in rows
        if any(cell is not None for cell in row)
    ]


def import_staff(*, organization, actor, rows: list[dict]) -> dict:
    created = 0
    errors = []
    for index, row in enumerate(rows, start=2):
        try:
            _import_one_row(organization=organization, actor=actor, row=row)
            created += 1
        except BulkImportError as exc:
            errors.append({"row": index, "error": str(exc)})
        except IntegrityError:
            errors.append({"row": index, "error": "a staff member with these values already exists"})
    return {"created": created, "errors": errors}


def _clean(row: dict, key: str) -> str:
    value = row.get(key)
    return str(value).strip() if value is not None else ""


def _import_one_row(*, organization, actor, row: dict) -> Staff:
    missing = [col for col in REQUIRED_COLUMNS if not _clean(row, col)]
    if missing:
        raise BulkImportError(f"missing required column(s): {', '.join(missing)}")

    school_code = _clean(row, "school_code")
    school = School.objects.filter(organization=organization, code=school_code).first()
    if school is None:
        raise BulkImportError(f"no school with code {school_code!r} in this organization")

    employment_status = _clean(row, "employment_status").lower() or "active"
    if employment_status not in _EMPLOYMENT_STATUSES:
        raise BulkImportError(f"invalid employment_status {employment_status!r}")

    fields = {
        "first_name": _clean(row, "first_name"),
        "last_name": _clean(row, "last_name"),
        "position": _clean(row, "position"),
        "date_joined": _clean(row, "date_joined"),
        "employment_status": employment_status,
        "phone": _clean(row, "phone"),
        "email": _clean(row, "email"),
    }
    employee_number = _clean(row, "employee_number")
    if employee_number:
        fields["employee_number"] = employee_number

    with transaction.atomic():
        return staff_service.create_staff(school=school, actor=actor, **fields)
