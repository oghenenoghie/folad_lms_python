"""CSV/XLSX bulk import for Student records — an admin onboarding a new
school's roster in one upload instead of one manual form submission per
student. Each row is created independently (its own savepoint): a bad row
is reported and skipped rather than aborting the whole batch, matching
apps.report_cards' bulk-export partial-success shape.
"""
import csv
import io

from django.db import IntegrityError, transaction

from apps.schools.models import School
from apps.students.models import ENROLLMENT_STATUS_CHOICES, GENDER_CHOICES, Student

from . import student_service

REQUIRED_COLUMNS = ["school_code", "first_name", "last_name", "date_of_birth"]
_ENROLLMENT_STATUSES = {choice for choice, _ in ENROLLMENT_STATUS_CHOICES}
_GENDERS = {choice for choice, _ in GENDER_CHOICES}


class BulkImportError(ValueError):
    pass


def parse_rows(*, filename: str, content: bytes) -> list[dict]:
    """One dict per data row, keyed by the header row's column names.
    Supports .csv (comma-separated, UTF-8) and .xlsx (first sheet)."""
    if filename.lower().endswith(".xlsx"):
        return _parse_xlsx(content)
    return _parse_csv(content)


def _parse_csv(content: bytes) -> list[dict]:
    # utf-8-sig strips a leading BOM, which Excel adds when it "Save As"s a
    # CSV — without this, the first header's name would come through with
    # a stray ﻿ prefix and never match REQUIRED_COLUMNS.
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows, [])]
    return [
        {header[i]: cell for i, cell in enumerate(row) if i < len(header)}
        for row in rows
        if any(cell is not None for cell in row)
    ]


def import_students(*, organization, actor, rows: list[dict]) -> dict:
    """Returns {"created": N, "errors": [{"row": <1-based, header=row 1>,
    "error": str}]}."""
    created = 0
    errors = []
    for index, row in enumerate(rows, start=2):
        try:
            _import_one_row(organization=organization, actor=actor, row=row)
            created += 1
        except BulkImportError as exc:
            errors.append({"row": index, "error": str(exc)})
        except IntegrityError:
            errors.append({"row": index, "error": "a student with these values already exists"})
    return {"created": created, "errors": errors}


def _clean(row: dict, key: str) -> str:
    value = row.get(key)
    return str(value).strip() if value is not None else ""


def _import_one_row(*, organization, actor, row: dict) -> Student:
    missing = [col for col in REQUIRED_COLUMNS if not _clean(row, col)]
    if missing:
        raise BulkImportError(f"missing required column(s): {', '.join(missing)}")

    school_code = _clean(row, "school_code")
    school = School.objects.filter(organization=organization, code=school_code).first()
    if school is None:
        raise BulkImportError(f"no school with code {school_code!r} in this organization")

    gender = _clean(row, "gender").lower()
    if gender and gender not in _GENDERS:
        raise BulkImportError(f"invalid gender {gender!r}")

    enrollment_status = _clean(row, "enrollment_status").lower() or "active"
    if enrollment_status not in _ENROLLMENT_STATUSES:
        raise BulkImportError(f"invalid enrollment_status {enrollment_status!r}")

    fields = {
        "first_name": _clean(row, "first_name"),
        "last_name": _clean(row, "last_name"),
        "date_of_birth": _clean(row, "date_of_birth"),
        "email": _clean(row, "email"),
        "gender": gender,
        "enrollment_status": enrollment_status,
    }
    admission_number = _clean(row, "admission_number")
    if admission_number:
        fields["admission_number"] = admission_number

    with transaction.atomic():
        return student_service.create_student(school=school, actor=actor, **fields)
